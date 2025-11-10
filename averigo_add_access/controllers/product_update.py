import json
from openpyxl import load_workbook


from odoo.http import Controller, request, route


class ProductCostData(Controller):

    @route('/product_cost_update', type='http', auth='public', csrf=False)
    def product_cost_update_new(self):
        print(11111)
        # file_path = '/opt/odoo13/SmithVendingProductMasterReport.xlsx'
        # workbook = load_workbook(filename=file_path)
        # sheet = workbook.active
        # start_collecting = False
        # data = []
        #
        # for row in sheet.iter_rows(values_only=True):
        #     if start_collecting:
        #         data.append(row)
        #     elif row and 'Product Code' in row:
        #         start_collecting = True
        #         data.append(row)
        # data_without_header = data[1:]
        # for data in data_without_header:
        #     code = data[0]
        #     product_cost = data[9]
        #     product_id = request.env['product.product'].sudo().search([('default_code','=', code),('company_id','=',142)])
        #     product_id.with_context(
        #         force_company=142).sudo().write(
        #         {
        #             'standard_price': product_cost})
